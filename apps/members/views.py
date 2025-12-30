# members/views.py

"""
Member Management Views

Comprehensive view functions for:
- Member Registration and Profile Management (using Wizard)
- Payment Methods and Financial Information
- Next of Kin and Emergency Contacts
- Member Groups and Group Memberships
- Additional Contacts and Communications
- Reports and Analytics

All views delegate business logic to services.py
Uses stats.py for comprehensive statistics and analytics
Uses SweetAlert2 for all notifications via Django messages
Preserves SessionWizardView for member registration
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, Prefetch
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.files.storage import FileSystemStorage
from formtools.wizard.views import SessionWizardView
from datetime import timedelta, date, datetime
from decimal import Decimal
import os
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO

from .models import (
    Member,
    MemberPaymentMethod,
    NextOfKin,
    MemberAdditionalContact,
    MemberGroup,
    GroupMembership,
)

from .forms import (
    MEMBER_WIZARD_FORMS,
    MEMBER_WIZARD_STEP_NAMES,
    MemberForm,
    MemberPaymentMethodForm,
    NextOfKinForm,
    MemberAdditionalContactForm,
    MemberGroupForm,
    GroupMembershipForm,
    MemberFilterForm,
    MemberPaymentMethodFilterForm,
    NextOfKinFilterForm,
    MemberGroupFilterForm,
    GroupMembershipFilterForm,
)

# Import services
from .services import (
    MemberService,
    KYCService,
    PaymentMethodService,
    NextOfKinService,
    GroupMembershipService,
)

# Import stats functions
from . import stats as member_stats

from core.utils import format_money

logger = logging.getLogger(__name__)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def members_dashboard(request):
    """Main members dashboard with overview statistics - USES stats.py"""
    
    try:
        # Use comprehensive overview from stats.py
        overview = member_stats.get_dashboard_summary()
        
        # Get additional statistics
        today = timezone.now().date()
        thirty_days_ago = today - timedelta(days=30)
        
        member_stats_data = member_stats.get_member_statistics({
            'date_range': (thirty_days_ago, today)
        })
        
        payment_method_stats = member_stats.get_payment_method_statistics()
        nok_stats = member_stats.get_next_of_kin_statistics()
        group_stats = member_stats.get_group_statistics()
        
    except Exception as e:
        logger.error(f"Error getting dashboard statistics: {e}")
        overview = {}
        member_stats_data = {}
        payment_method_stats = {}
        nok_stats = {}
        group_stats = {}
    
    # Get recent activities (limited queries for display)
    recent_members = Member.objects.order_by('-created_at')[:10]
    
    pending_approvals = Member.objects.filter(
        status='PENDING_APPROVAL'
    ).order_by('created_at')[:10]
    
    recent_groups = MemberGroup.objects.select_related(
        'group_leader'
    ).order_by('-created_at')[:10]
    
    # Get members needing attention
    kyc_expired = Member.objects.filter(
        Q(kyc_status='EXPIRED') | Q(kyc_status='REQUIRES_UPDATE')
    ).order_by('-updated_at')[:10]
    
    high_risk_members = Member.objects.filter(
        risk_rating__in=['HIGH', 'VERY_HIGH']
    ).order_by('-credit_score')[:10]
    
    # Get incomplete profiles
    incomplete_profiles = Member.objects.filter(
        Q(phone_primary__isnull=True) |
        Q(physical_address__isnull=True) |
        Q(employment_status__isnull=True)
    ).order_by('-created_at')[:10]
    
    context = {
        'overview': overview,
        'member_stats': member_stats_data,
        'payment_method_stats': payment_method_stats,
        'nok_stats': nok_stats,
        'group_stats': group_stats,
        'recent_members': recent_members,
        'pending_approvals': pending_approvals,
        'recent_groups': recent_groups,
        'kyc_expired': kyc_expired,
        'high_risk_members': high_risk_members,
        'incomplete_profiles': incomplete_profiles,
    }
    
    return render(request, 'members/dashboard.html', context)


# =============================================================================
# MEMBER VIEWS
# =============================================================================

@login_required
def member_list(request):
    """List all members - HTMX loads data on page load"""
    
    # Initialize filter form
    filter_form = MemberFilterForm()
    
    # Get initial stats from stats.py
    try:
        initial_stats = member_stats.get_member_statistics()
    except Exception as e:
        logger.error(f"Error getting member statistics: {e}")
        initial_stats = {}
    
    context = {
        'filter_form': filter_form,
        'stats': initial_stats,
        'Member': Member,
    }
    
    return render(request, 'members/list.html', context)

@login_required
def member_print_view(request):
    """Generate printable member list with selected fields"""
    
    # Get selected fields from the modal
    selected_fields = request.GET.getlist('fields')
    if not selected_fields:
        # Default fields if none selected
        selected_fields = ['member_number', 'full_name', 'phone_primary', 'member_category', 'status', 'credit_score', 'kyc_status']
    
    # Get additional options
    include_stats = request.GET.get('include_stats') == 'true'
    landscape = request.GET.get('landscape') == 'true'
    
    # Get filter parameters from URL
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    member_category = request.GET.get('member_category', '')
    membership_plan = request.GET.get('membership_plan', '')
    gender = request.GET.get('gender', '')
    kyc_status = request.GET.get('kyc_status', '')
    risk_rating = request.GET.get('risk_rating', '')
    employment_status = request.GET.get('employment_status', '')
    
    # Build queryset
    members = Member.objects.select_related().order_by('member_number')
    
    # Apply filters (same as member_search)
    if query:
        members = members.filter(
            Q(member_number__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(phone_primary__icontains=query) |
            Q(personal_email__icontains=query)
        )
    
    if status:
        members = members.filter(status=status)
    
    if member_category:
        members = members.filter(member_category=member_category)
    
    if membership_plan:
        members = members.filter(membership_plan=membership_plan)
    
    if gender:
        members = members.filter(gender=gender)
    
    if kyc_status:
        members = members.filter(kyc_status=kyc_status)
    
    if risk_rating:
        members = members.filter(risk_rating=risk_rating)
    
    if employment_status:
        members = members.filter(employment_status=employment_status)
    
    # Calculate stats only if requested
    stats = None
    if include_stats:
        total = members.count()
        active_count = members.filter(status='ACTIVE').count()
        
        stats = {
            'total': total,
            'active': active_count,
            'active_percentage': round((active_count / total * 100), 1) if total > 0 else 0,
            'pending_approval': members.filter(status='PENDING_APPROVAL').count(),
            'avg_credit_score': round(members.aggregate(Avg('credit_score'))['credit_score__avg'] or 0),
        }
    
    # Field display names mapping
    field_names = {
        'member_number': 'Member Number',
        'full_name': 'Full Name',
        'id_number': 'ID Number',
        'date_of_birth': 'Date of Birth',
        'gender': 'Gender',
        'nationality': 'Nationality',
        'phone_primary': 'Primary Phone',
        'phone_secondary': 'Secondary Phone',
        'personal_email': 'Email',
        'postal_address': 'Postal Address',
        'physical_address': 'Physical Address',
        'member_category': 'Category',
        'membership_plan': 'Plan',
        'status': 'Status',
        'membership_date': 'Membership Date',
        'branch': 'Branch',
        'credit_score': 'Credit Score',
        'kyc_status': 'KYC Status',
        'risk_rating': 'Risk Rating',
        'monthly_income': 'Monthly Income',
        'employment_status': 'Employment',
    }
    
    # Create ordered list of field display names for template
    selected_field_names = [field_names.get(field, field.replace('_', ' ').title()) for field in selected_fields]
    
    context = {
        'members': members,
        'stats': stats,
        'now': timezone.now(),
        'selected_fields': selected_fields,
        'selected_field_names': selected_field_names,  # ✅ Add this for proper headers
        'field_names': field_names,
        'landscape': landscape,
    }
    
    return render(request, 'members/print.html', context)  

# =============================================================================
# MEMBER WIZARD FOR CREATION - UPDATED
# =============================================================================

class MemberWizardFileStorage(FileSystemStorage):
    """Custom storage for handling file uploads in wizard"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.location = os.path.join(self.location, 'wizard_temp')


class MemberCreateWizard(SessionWizardView):
    """
    Multi-step wizard for creating a member - DELEGATES TO MemberService
    
    Steps:
    1. Basic Information - personal details and identification
    2. Contact Information - address and contact details
    3. Employment Information - employment and income details
    4. Membership Information - membership category and plan
    5. Next of Kin - primary next of kin (optional)
    6. Confirmation - review and confirm
    
    Note: Member number is automatically generated by pre_save signal in signals.py
    """

    form_list = MEMBER_WIZARD_FORMS
    template_name = 'members/wizard.html'  
    file_storage = MemberWizardFileStorage()

    def get_template_names(self):
        """Return the template for all steps"""
        return [self.template_name]

    def get_context_data(self, form, **kwargs):
        """Add step names and progress tracking"""
        context = super().get_context_data(form=form, **kwargs)

        total_steps = len(self.form_list)
        current_step_index = list(self.form_list).index(self.steps.current)

        context.update({
            'step_names': MEMBER_WIZARD_STEP_NAMES,
            'current_step_name': MEMBER_WIZARD_STEP_NAMES.get(
                self.steps.current, 'Step'
            ),
            'progress_percentage': ((current_step_index) / (total_steps - 1)) * 100 if total_steps > 1 else 100,
        })

        # Add review data for confirmation step
        if self.steps.current == 'confirmation':
            context['basic_data'] = self.get_cleaned_data_for_step('basic_info')
            context['contact_data'] = self.get_cleaned_data_for_step('contact_info')
            context['employment_data'] = self.get_cleaned_data_for_step('employment_info')
            context['membership_data'] = self.get_cleaned_data_for_step('membership_info')
            context['nok_data'] = self.get_cleaned_data_for_step('next_of_kin')

        return context

    def get_form_kwargs(self, step=None):
        """Pass additional kwargs to forms if needed"""
        kwargs = super().get_form_kwargs(step)
        
        # You can add additional form kwargs here if needed
        # For example, passing the current user:
        # kwargs['user'] = self.request.user
        
        return kwargs

    @transaction.atomic
    def done(self, form_list, **kwargs):
        """
        Persist all wizard data and create member.
        DELEGATES business logic to MemberService.register_member()
        
        Member number generation is handled automatically by the pre_save signal
        in signals.py (generate_member_number function).
        """
        
        logger.info("=" * 80)
        logger.info("WIZARD DONE - Delegating to MemberService")
        logger.info("=" * 80)

        try:
            # Merge cleaned data from all steps
            form_data = {}
            form_dict = {}
            
            for step, form in zip(self.form_list.keys(), form_list):
                form_data.update(form.cleaned_data)
                form_dict[step] = form

            # ------------------------------------------------------------------
            # Prepare member data for service
            # ------------------------------------------------------------------
            member_data = {
                # Basic info
                'id_number': form_data.get('id_number'),
                'id_type': form_data.get('id_type'),
                'title': form_data.get('title'),
                'first_name': form_data.get('first_name'),
                'last_name': form_data.get('last_name'),
                'middle_name': form_data.get('middle_name'),
                'date_of_birth': form_data.get('date_of_birth'),
                'place_of_birth': form_data.get('place_of_birth'),
                'gender': form_data.get('gender'),
                'marital_status': form_data.get('marital_status'),
                'nationality': form_data.get('nationality'),
                'religious_affiliation': form_data.get('religious_affiliation'),
                
                # Contact info
                'personal_email': form_data.get('personal_email', ''),
                'phone_primary': form_data.get('phone_primary'),
                'physical_address': form_data.get('physical_address'),
                'postal_address': form_data.get('postal_address', ''),
                'postal_code': form_data.get('postal_code', ''),
                'city': form_data.get('city', ''),
                'state_province': form_data.get('state_province', ''),
                'country': form_data.get('country', 'UG'),
                
                # Employment info
                'employment_status': form_data.get('employment_status'),
                'occupation': form_data.get('occupation', ''),
                'employer': form_data.get('employer', ''),
                'employer_address': form_data.get('employer_address', ''),
                'monthly_income': form_data.get('monthly_income'),
                'income_source': form_data.get('income_source', ''),
                'other_income_sources': form_data.get('other_income_sources', ''),
                
                # Membership info
                'member_category': form_data.get('member_category', 'REGULAR'),
                'membership_plan': form_data.get('membership_plan', 'BASIC'),
                'membership_date': form_data.get('membership_date'),
                'membership_application_date': form_data.get('membership_application_date'),
                'status': form_data.get('status', 'PENDING_APPROVAL'),
                'tax_id': form_data.get('tax_id', ''),
            }
            
            # Note: Member number is NOT passed here - it will be auto-generated
            # by the pre_save signal in signals.py
            # The user's choice (auto_generate_member_number) is ignored since
            # all member numbers are auto-generated for consistency

            # ------------------------------------------------------------------
            # DELEGATE to MemberService
            # ------------------------------------------------------------------
            success, result = MemberService.register_member(
                member_data=member_data,
                created_by=self.request.user
            )
            
            if not success:
                messages.error(
                    self.request,
                    f"Failed to register member: {result}",
                    extra_tags='sweetalert-error'
                )
                return redirect('members:member_list')
            
            member = result

            # ------------------------------------------------------------------
            # Handle Next of Kin - DELEGATE to NextOfKinService
            # ------------------------------------------------------------------
            nok_option = form_data.get('nok_option', 'skip')
            
            if nok_option == 'new':
                nok_success, nok_result = NextOfKinService.add_next_of_kin(
                    member=member,
                    name=form_data.get('nok_name'),
                    relation=form_data.get('nok_relation'),
                    contact=form_data.get('nok_contact'),
                    email=form_data.get('nok_email', ''),
                    address=form_data.get('nok_address', ''),
                    id_number=form_data.get('nok_id_number', ''),
                    is_primary=True,
                    is_emergency_contact=form_data.get('nok_is_emergency_contact', True)
                )
                
                if not nok_success:
                    logger.warning(f"Failed to add next of kin: {nok_result}")

            # ------------------------------------------------------------------
            # Success - Member number was auto-generated by signal
            # ------------------------------------------------------------------
            messages.success(
                self.request,
                f"Member {member.get_full_name()} "
                f"(#{member.member_number}) was created successfully!",
                extra_tags='sweetalert'
            )

            return redirect('members:member_profile', pk=member.pk)

        except Exception as exc:
            logger.exception("Error in wizard done method:")
            logger.exception(exc)
            
            messages.error(
                self.request,
                f"Error creating member: {exc}",
                extra_tags='sweetalert-error'
            )
            return redirect('members:member_list')


# View entry point
member_create = MemberCreateWizard.as_view()

@login_required
def member_edit(request, pk):
    """Edit existing member - DELEGATES to MemberService"""
    member = get_object_or_404(Member, pk=pk)

    if request.method == "POST":
        form = MemberForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            # Get updated data
            update_data = form.cleaned_data
            
            # DELEGATE to MemberService
            success, message = MemberService.update_member_profile(
                member=member,
                update_data=update_data,
                updated_by=request.user
            )
            
            if success:
                messages.success(
                    request,
                    f"Member {member.get_full_name()} was updated successfully",
                    extra_tags='sweetalert'
                )
                return redirect("members:member_profile", pk=member.pk)
            else:
                messages.error(
                    request,
                    message,
                    extra_tags='sweetalert-error'
                )
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = MemberForm(instance=member)

    context = {
        'form': form,
        'member': member,
        'title': 'Update Member',
    }

    return render(request, 'members/form.html', context)


@login_required
def member_profile(request, pk):
    """View member profile with all related information - USES stats.py"""
    member = get_object_or_404(
        Member.objects.prefetch_related(
            Prefetch(
                'payment_methods',
                queryset=MemberPaymentMethod.objects.order_by('-is_primary', 'provider')
            ),
            Prefetch(
                'next_of_kin',
                queryset=NextOfKin.objects.order_by('-is_primary', 'name')
            ),
            Prefetch(
                'additional_contacts',
                queryset=MemberAdditionalContact.objects.order_by('contact_type')
            ),
            Prefetch(
                'groupmembership_set',  # ✅ CHANGED FROM 'group_memberships'
                queryset=GroupMembership.objects.filter(is_active=True).select_related('group')
            )
        ),
        pk=pk
    )

    # Get member summary from stats.py
    try:
        summary = {
            'member_number': member.member_number,
            'full_name': member.get_full_name(),
            'status': member.get_status_display(),
            'age': member.age,
            'membership_duration_years': round(member.membership_duration_years, 1),
            'credit_score': member.credit_score,
            'risk_rating': member.get_risk_rating_display(),
            'monthly_income': str(member.monthly_income) if member.monthly_income else None,
            'monthly_income_formatted': format_money(member.monthly_income) if member.monthly_income else None,
            'kyc_status': member.get_kyc_status_display(),
            'kyc_verified': member.is_kyc_verified,
            
            # Related counts
            'savings_accounts': member.savings_accounts.count() if hasattr(member, 'savings_accounts') else 0,
            'active_loans': member.get_active_loans_count() if hasattr(member, 'get_active_loans_count') else 0,
            'payment_methods': member.payment_methods.filter(is_active=True).count(),
            'next_of_kin': member.next_of_kin.count(),
            'group_memberships': member.groupmembership_set.filter(is_active=True).count(),  # ✅ CHANGED
            
            # Financial aggregates
            'total_savings': str(member.get_total_savings()) if hasattr(member, 'get_total_savings') else '0.00',
            'total_loans': str(member.get_total_loans()) if hasattr(member, 'get_total_loans') else '0.00',
            'total_dividends': str(member.get_total_dividends()) if hasattr(member, 'get_total_dividends') else '0.00',
        }
    except Exception as e:
        logger.error(f"Error getting member summary: {e}")
        summary = {}

    # Get related data
    payment_methods = member.payment_methods.all()
    primary_payment = payment_methods.filter(is_primary=True).first()
    next_of_kin = member.next_of_kin.all()
    primary_nok = next_of_kin.filter(is_primary=True).first()
    emergency_contact = next_of_kin.filter(is_emergency_contact=True).first()
    additional_contacts = member.additional_contacts.all()
    group_memberships = member.groupmembership_set.filter(is_active=True)  # ✅ CHANGED

    context = {
        'member': member,
        'summary': summary,
        'payment_methods': payment_methods,
        'primary_payment': primary_payment,
        'next_of_kin': next_of_kin,
        'primary_nok': primary_nok,
        'emergency_contact': emergency_contact,
        'additional_contacts': additional_contacts,
        'group_memberships': group_memberships,
    }
    
    return render(request, "members/profile.html", context)


@login_required
def member_activate(request, pk):
    """Activate a member - DELEGATES to MemberService"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        # DELEGATE to MemberService
        success, message = MemberService.approve_member(
            member=member,
            approved_by=request.user
        )
        
        if success:
            messages.success(
                request,
                f"Member {member.get_full_name()} has been activated successfully",
                extra_tags='sweetalert'
            )
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=member.pk)
    
    return redirect('members:member_profile', pk=member.pk)


@login_required
def member_suspend(request, pk):
    """Suspend a member - DELEGATES to MemberService"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'No reason provided')
        
        # DELEGATE to MemberService
        success, message = MemberService.suspend_member(
            member=member,
            reason=reason,
            suspended_by=request.user
        )
        
        if success:
            messages.warning(
                request,
                f"Member {member.get_full_name()} has been suspended",
                extra_tags='sweetalert'
            )
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=member.pk)
    
    return redirect('members:member_profile', pk=member.pk)


@login_required
def member_reactivate(request, pk):
    """Reactivate a suspended member - DELEGATES to MemberService"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        # DELEGATE to MemberService
        success, message = MemberService.reactivate_member(
            member=member,
            reactivated_by=request.user
        )
        
        if success:
            messages.success(
                request,
                f"Member {member.get_full_name()} has been reactivated",
                extra_tags='sweetalert'
            )
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=member.pk)
    
    return redirect('members:member_profile', pk=member.pk)


@login_required
def member_update_credit_score(request, pk):
    """Update member's credit score - DELEGATES to MemberService"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        # DELEGATE to MemberService
        success, new_score, message = MemberService.update_credit_score(member)
        
        if success:
            messages.success(
                request,
                f"Credit score updated to {new_score}",
                extra_tags='sweetalert'
            )
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=member.pk)
    
    return redirect('members:member_profile', pk=member.pk)


# =============================================================================
# KYC VIEWS
# =============================================================================

@login_required
def member_verify_kyc(request, pk):
    """Verify member's KYC - DELEGATES to KYCService"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        expiry_years = int(request.POST.get('expiry_years', 2))
        
        # DELEGATE to KYCService
        success, message = KYCService.verify_kyc(
            member=member,
            expiry_years=expiry_years,
            verified_by=request.user
        )
        
        if success:
            messages.success(request, message, extra_tags='sweetalert')
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=member.pk)
    
    return redirect('members:member_profile', pk=member.pk)


@login_required
def member_reject_kyc(request, pk):
    """Reject member's KYC - DELEGATES to KYCService"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'No reason provided')
        
        # DELEGATE to KYCService
        success, message = KYCService.reject_kyc(
            member=member,
            reason=reason,
            rejected_by=request.user
        )
        
        if success:
            messages.warning(request, message, extra_tags='sweetalert')
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=member.pk)
    
    return redirect('members:member_profile', pk=member.pk)


# =============================================================================
# PAYMENT METHOD VIEWS
# =============================================================================

@login_required
def payment_method_list(request):
    """List all payment methods - HTMX loads data on page load"""
    
    filter_form = MemberPaymentMethodFilterForm()
    
    try:
        initial_stats = member_stats.get_payment_method_statistics()
    except Exception as e:
        logger.error(f"Error getting payment method statistics: {e}")
        initial_stats = {}
    
    context = {
        'filter_form': filter_form,
        'stats': initial_stats,
    }
    
    return render(request, 'members/payment_methods/list.html', context)


@login_required
def payment_method_create(request, member_pk):
    """Add a new payment method - DELEGATES to PaymentMethodService"""
    member = get_object_or_404(Member, pk=member_pk)
    
    if request.method == 'POST':
        form = MemberPaymentMethodForm(request.POST, member=member)
        if form.is_valid():
            # DELEGATE to PaymentMethodService
            success, result = PaymentMethodService.add_payment_method(
                member=member,
                method_type=form.cleaned_data['method_type'],
                provider=form.cleaned_data['provider'],
                account_number=form.cleaned_data['account_number'],
                account_name=form.cleaned_data['account_name'],
                is_primary=form.cleaned_data.get('is_primary', False),
                account_type=form.cleaned_data.get('account_type'),
                branch=form.cleaned_data.get('branch'),
                notes=form.cleaned_data.get('notes')
            )
            
            if success:
                messages.success(
                    request,
                    f"Payment method added for {member.get_full_name()}",
                    extra_tags='sweetalert'
                )
                return redirect('members:member_profile', pk=member.pk)
            else:
                messages.error(request, result, extra_tags='sweetalert-error')
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = MemberPaymentMethodForm(member=member)
    
    context = {
        'form': form,
        'member': member,
        'title': f'Add Payment Method for {member.get_full_name()}',
    }
    return render(request, 'members/payment_methods/form.html', context)


@login_required
def payment_method_edit(request, pk):
    """Edit payment method"""
    payment_method = get_object_or_404(MemberPaymentMethod, pk=pk)
    
    if request.method == 'POST':
        form = MemberPaymentMethodForm(
            request.POST, 
            instance=payment_method, 
            member=payment_method.member
        )
        if form.is_valid():
            payment_method = form.save()
            messages.success(
                request,
                "Payment method updated successfully",
                extra_tags='sweetalert'
            )
            return redirect('members:member_profile', pk=payment_method.member.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = MemberPaymentMethodForm(instance=payment_method, member=payment_method.member)
    
    context = {
        'form': form,
        'payment_method': payment_method,
        'member': payment_method.member,
        'title': 'Edit Payment Method',
    }
    return render(request, 'members/payment_methods/form.html', context)


@login_required
def payment_method_verify(request, pk):
    """Verify payment method - DELEGATES to PaymentMethodService"""
    payment_method = get_object_or_404(MemberPaymentMethod, pk=pk)
    
    if request.method == 'POST':
        # DELEGATE to PaymentMethodService
        success, message = PaymentMethodService.verify_payment_method(
            payment_method=payment_method,
            verified_by=request.user
        )
        
        if success:
            messages.success(request, message, extra_tags='sweetalert')
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=payment_method.member.pk)
    
    return redirect('members:member_profile', pk=payment_method.member.pk)


@login_required
def payment_method_set_primary(request, pk):
    """Set payment method as primary - DELEGATES to PaymentMethodService"""
    payment_method = get_object_or_404(MemberPaymentMethod, pk=pk)
    
    if request.method == 'POST':
        # DELEGATE to PaymentMethodService
        success, message = PaymentMethodService.set_primary_payment_method(payment_method)
        
        if success:
            messages.success(request, message, extra_tags='sweetalert')
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:member_profile', pk=payment_method.member.pk)
    
    return redirect('members:member_profile', pk=payment_method.member.pk)


# =============================================================================
# NEXT OF KIN VIEWS
# =============================================================================

@login_required
def next_of_kin_list(request):
    """List all next of kin - HTMX loads data on page load"""
    
    filter_form = NextOfKinFilterForm()
    
    try:
        initial_stats = member_stats.get_next_of_kin_statistics()
    except Exception as e:
        logger.error(f"Error getting next of kin statistics: {e}")
        initial_stats = {}
    
    context = {
        'filter_form': filter_form,
        'stats': initial_stats,
    }
    
    return render(request, 'members/next_of_kin/list.html', context)


@login_required
def next_of_kin_create(request, member_pk):
    """Add a new next of kin - DELEGATES to NextOfKinService"""
    member = get_object_or_404(Member, pk=member_pk)
    
    if request.method == 'POST':
        form = NextOfKinForm(request.POST, member=member)
        if form.is_valid():
            # DELEGATE to NextOfKinService
            success, result = NextOfKinService.add_next_of_kin(
                member=member,
                name=form.cleaned_data['name'],
                relation=form.cleaned_data['relation'],
                contact=form.cleaned_data['contact'],
                is_primary=form.cleaned_data.get('is_primary', False),
                is_beneficiary=form.cleaned_data.get('is_beneficiary', False),
                beneficiary_percentage=form.cleaned_data.get('beneficiary_percentage', Decimal('0.00')),
                email=form.cleaned_data.get('email'),
                address=form.cleaned_data.get('address'),
                id_number=form.cleaned_data.get('id_number'),
                date_of_birth=form.cleaned_data.get('date_of_birth'),
                is_emergency_contact=form.cleaned_data.get('is_emergency_contact', False)
            )
            
            if success:
                messages.success(
                    request,
                    f"Next of kin added for {member.get_full_name()}",
                    extra_tags='sweetalert'
                )
                return redirect('members:member_profile', pk=member.pk)
            else:
                messages.error(request, result, extra_tags='sweetalert-error')
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = NextOfKinForm(member=member)
    
    context = {
        'form': form,
        'member': member,
        'title': f'Add Next of Kin for {member.get_full_name()}',
    }
    return render(request, 'members/next_of_kin/form.html', context)


@login_required
def next_of_kin_edit(request, pk):
    """Edit next of kin - DELEGATES to NextOfKinService"""
    nok = get_object_or_404(NextOfKin, pk=pk)
    
    if request.method == 'POST':
        form = NextOfKinForm(request.POST, instance=nok, member=nok.member)
        if form.is_valid():
            # DELEGATE to NextOfKinService
            success, message = NextOfKinService.update_next_of_kin(
                next_of_kin=nok,
                update_data=form.cleaned_data,
                updated_by=request.user
            )
            
            if success:
                messages.success(
                    request,
                    "Next of kin updated successfully",
                    extra_tags='sweetalert'
                )
                return redirect('members:member_profile', pk=nok.member.pk)
            else:
                messages.error(request, message, extra_tags='sweetalert-error')
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = NextOfKinForm(instance=nok, member=nok.member)
    
    context = {
        'form': form,
        'nok': nok,
        'member': nok.member,
        'title': 'Edit Next of Kin',
    }
    return render(request, 'members/next_of_kin/form.html', context)


@login_required
def next_of_kin_set_primary(request, pk):
    """Set next of kin as primary"""
    nok = get_object_or_404(NextOfKin, pk=pk)
    
    if request.method == 'POST':
        nok.make_primary()
        messages.success(
            request,
            f"{nok.name} is now the primary next of kin",
            extra_tags='sweetalert'
        )
        return redirect('members:member_profile', pk=nok.member.pk)
    
    return redirect('members:member_profile', pk=nok.member.pk)


# =============================================================================
# ADDITIONAL CONTACT VIEWS
# =============================================================================

@login_required
def additional_contact_create(request, member_pk):
    """Add additional contact to a member"""
    member = get_object_or_404(Member, pk=member_pk)
    
    if request.method == 'POST':
        form = MemberAdditionalContactForm(request.POST, member=member)
        if form.is_valid():
            contact = form.save(commit=False)
            contact.member = member
            contact.save()
            messages.success(
                request,
                f"Additional contact added for {member.get_full_name()}",
                extra_tags='sweetalert'
            )
            return redirect('members:member_profile', pk=member.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = MemberAdditionalContactForm(member=member)
    
    context = {
        'form': form,
        'member': member,
        'title': f'Add Additional Contact for {member.get_full_name()}',
    }
    return render(request, 'members/additional_contacts/form.html', context)


@login_required
def additional_contact_edit(request, pk):
    """Edit additional contact"""
    contact = get_object_or_404(MemberAdditionalContact, pk=pk)
    
    if request.method == 'POST':
        form = MemberAdditionalContactForm(
            request.POST, 
            instance=contact, 
            member=contact.member
        )
        if form.is_valid():
            contact = form.save()
            messages.success(
                request,
                "Additional contact updated successfully",
                extra_tags='sweetalert'
            )
            return redirect('members:member_profile', pk=contact.member.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = MemberAdditionalContactForm(instance=contact, member=contact.member)
    
    context = {
        'form': form,
        'contact': contact,
        'member': contact.member,
        'title': 'Edit Additional Contact',
    }
    return render(request, 'members/additional_contacts/form.html', context)


# =============================================================================
# MEMBER GROUP VIEWS
# =============================================================================

@login_required
def group_list(request):
    """List all member groups - HTMX loads data on page load"""
    
    filter_form = MemberGroupFilterForm()
    
    try:
        initial_stats = member_stats.get_group_statistics()
    except Exception as e:
        logger.error(f"Error getting group statistics: {e}")
        initial_stats = {}
    
    context = {
        'filter_form': filter_form,
        'stats': initial_stats,
    }
    
    return render(request, 'members/groups/list.html', context)


@login_required
def group_create(request):
    """Create a new member group"""
    if request.method == 'POST':
        form = MemberGroupForm(request.POST)
        if form.is_valid():
            group = form.save()
            messages.success(
                request,
                f"Group {group.name} created successfully",
                extra_tags='sweetalert'
            )
            return redirect('members:group_detail', pk=group.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = MemberGroupForm()

    context = {
        'form': form,
        'title': 'Create New Group',
    }
    return render(request, 'members/groups/form.html', context)


@login_required
def group_edit(request, pk):
    """Edit member group"""
    group = get_object_or_404(MemberGroup, pk=pk)
    
    if request.method == 'POST':
        form = MemberGroupForm(request.POST, instance=group)
        if form.is_valid():
            group = form.save()
            messages.success(
                request,
                f"Group {group.name} updated successfully",
                extra_tags='sweetalert'
            )
            return redirect('members:group_detail', pk=group.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = MemberGroupForm(instance=group)
    
    context = {
        'form': form,
        'group': group,
        'title': 'Edit Group',
    }
    return render(request, 'members/groups/form.html', context)


@login_required
def group_detail(request, pk):
    """View group details - USES stats.py"""
    group = get_object_or_404(
        MemberGroup.objects.select_related(
            'group_leader', 'group_secretary', 'group_treasurer'
        ).prefetch_related(
            Prefetch(
                'groupmembership_set',
                queryset=GroupMembership.objects.filter(is_active=True).select_related('member')
            )
        ),
        pk=pk
    )

    try:
        aggregates = group.groupmembership_set.filter(is_active=True).aggregate(
            total_contributions=Sum('total_contributions'),
            avg_contributions=Avg('total_contributions'),
            avg_attendance=Avg('meeting_attendance_rate')
        )
        
        group_summary = {
            'name': group.name,
            'group_type': group.get_group_type_display(),
            'is_active': group.is_active,
            'is_full': group.is_full,
            'formation_date': str(group.formation_date),
            'group_age_days': group.group_age_days,
            'member_count': group.member_count,
            'available_slots': group.available_slots,
            'total_contributions': aggregates['total_contributions'] or Decimal('0.00'),
            'avg_contributions': aggregates['avg_contributions'] or Decimal('0.00'),
            'avg_attendance': round(aggregates['avg_attendance'] or 0, 1),
        }
    except Exception as e:
        logger.error(f"Error getting group summary: {e}")
        group_summary = {}

    memberships = group.groupmembership_set.filter(is_active=True).order_by('role', 'join_date')

    context = {
        'group': group,
        'summary': group_summary,
        'memberships': memberships,
    }
    return render(request, 'members/groups/detail.html', context)


@login_required
def group_add_member(request, group_pk):
    """Add member to group - DELEGATES to GroupMembershipService"""
    group = get_object_or_404(MemberGroup, pk=group_pk)
    
    if request.method == 'POST':
        form = GroupMembershipForm(request.POST)
        if form.is_valid():
            # DELEGATE to GroupMembershipService
            success, result = GroupMembershipService.join_group(
                member=form.cleaned_data['member'],
                group=group,
                role=form.cleaned_data.get('role', 'MEMBER'),
                monthly_contribution=form.cleaned_data.get('monthly_contribution')
            )
            
            if success:
                messages.success(
                    request,
                    f"{result.member.get_full_name()} added to {group.name}",
                    extra_tags='sweetalert'
                )
                return redirect('members:group_detail', pk=group.pk)
            else:
                messages.error(request, result, extra_tags='sweetalert-error')
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = GroupMembershipForm(initial={'group': group})
        form.fields['group'].widget.attrs['disabled'] = True
    
    context = {
        'form': form,
        'group': group,
        'title': f'Add Member to {group.name}',
    }
    return render(request, 'members/groups/add_member.html', context)


@login_required
def group_membership_edit(request, pk):
    """Edit group membership details"""
    membership = get_object_or_404(GroupMembership, pk=pk)
    
    if request.method == 'POST':
        form = GroupMembershipForm(request.POST, instance=membership)
        if form.is_valid():
            membership = form.save()
            messages.success(
                request,
                "Membership updated successfully",
                extra_tags='sweetalert'
            )
            return redirect('members:group_detail', pk=membership.group.pk)
        else:
            messages.error(
                request,
                "Please correct the errors in the form",
                extra_tags='sweetalert-error'
            )
    else:
        form = GroupMembershipForm(instance=membership)
        form.fields['member'].widget.attrs['disabled'] = True
        form.fields['group'].widget.attrs['disabled'] = True
    
    context = {
        'form': form,
        'membership': membership,
        'title': 'Edit Group Membership',
    }
    return render(request, 'members/groups/edit_membership.html', context)


@login_required
def group_membership_leave(request, pk):
    """Remove member from group - DELEGATES to GroupMembershipService"""
    membership = get_object_or_404(GroupMembership, pk=pk)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        
        # DELEGATE to GroupMembershipService
        success, message = GroupMembershipService.leave_group(
            membership=membership,
            reason=reason,
            left_by=request.user
        )
        
        if success:
            messages.success(request, message, extra_tags='sweetalert')
        else:
            messages.error(request, message, extra_tags='sweetalert-error')
        
        return redirect('members:group_detail', pk=membership.group.pk)
    
    return redirect('members:group_detail', pk=membership.group.pk)


# =============================================================================
# EXPORT VIEWS (Keep as is - they don't need service delegation)
# =============================================================================

@login_required
def export_members_excel(request):
    """Export members to Excel with filters applied"""
    
    # Get filter parameters
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    member_category = request.GET.get('member_category', '')
    
    # Apply filters
    members = Member.objects.all().order_by('member_number')
    
    if query:
        members = members.filter(
            Q(member_number__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )
    
    if status:
        members = members.filter(status=status)
    if member_category:
        members = members.filter(member_category=member_category)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    headers = ['#', 'Member No.', 'Full Name', 'Gender', 'Status', 'Phone', 'Email']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for idx, member in enumerate(members, start=1):
        ws.append([
            idx,
            member.member_number,
            member.get_full_name(),
            member.get_gender_display(),
            member.get_status_display(),
            member.phone_primary or '',
            member.personal_email or ''
        ])
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"members_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def export_members_pdf(request):
    """Export members to PDF with filters applied"""
    
    # Get filter parameters from request
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    member_category = request.GET.get('member_category', '')
    gender = request.GET.get('gender', '')
    
    # Apply filters
    members = Member.objects.all().order_by('member_number')
    
    if query:
        members = members.filter(
            Q(member_number__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )
    
    if status:
        members = members.filter(status=status)
    if member_category:
        members = members.filter(member_category=member_category)
    if gender:
        members = members.filter(gender=gender)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=18,
    )
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER,
    )
    
    # Title
    title = Paragraph("Member Directory Report", title_style)
    elements.append(title)
    
    # Subtitle with filters
    filter_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if query:
        filter_text += f" | Search: {query}"
    if status:
        filter_text += f" | Status: {dict(Member.STATUS_CHOICES).get(status)}"
    
    subtitle = Paragraph(filter_text, subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [['#', 'Member No.', 'Full Name', 'Gender', 'Age', 'Category', 'Status', 'Credit Score']]
    
    for idx, member in enumerate(members, start=1):
        try:
            age = str(member.age)
        except:
            age = 'N/A'
        
        row = [
            str(idx),
            member.member_number,
            member.get_full_name()[:30],
            member.get_gender_display(),
            age,
            member.get_member_category_display()[:15],
            member.get_status_display()[:15],
            str(member.credit_score)
        ]
        data.append(row)
    
    # Create table
    table = Table(data, colWidths=[0.4*inch, 1*inch, 2*inch, 0.8*inch, 0.6*inch, 1.2*inch, 1*inch, 0.8*inch])
    
    # Table style
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # # column
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Age column
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),  # Credit score column
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(table)
    
    # Summary
    elements.append(Spacer(1, 0.3*inch))
    summary_text = f"<b>Total Members:</b> {members.count()}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF value
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"members_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response